"""
Servicio de planillas Excel para carga masiva de alumnos (Módulo 0).

Formato esperado (una fila por alumno, con encabezados en la primera fila):

    | nombre        | rut          | curso        |
    | Juan Pérez    | 21.345.678-9 | 1° Medio A   |
    | Ana Rojas     | 22.111.222-3 | 1° Medio A   |

Detalles de tolerancia (el profesor no debería pelear con el formato):
  - Los encabezados se detectan sin distinguir mayúsculas/tildes/espacios.
  - Se aceptan sinónimos: "rut" | "identificador" | "matricula"; "curso" | "grupo".
  - Las filas vacías se ignoran; los errores se reportan por número de fila
    en vez de abortar toda la importación.
"""
from __future__ import annotations

import io
import unicodedata

from openpyxl import Workbook, load_workbook


# Sinónimos aceptados por columna → nombre canónico.
_ALIAS = {
    "nombre": "nombre", "alumno": "nombre", "nombrealumno": "nombre",
    "estudiante": "nombre", "nombrecompleto": "nombre",
    "rut": "identificador", "run": "identificador", "identificador": "identificador",
    "matricula": "identificador", "id": "identificador", "cedula": "identificador",
    "curso": "curso", "grupo": "curso", "seccion": "curso", "nivel": "curso",
}


def _norm(v: object) -> str:
    """Normaliza un encabezado: minúsculas, sin tildes, sin espacios/símbolos."""
    s = str(v or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(c for c in s if c.isalnum())


def parse_students_workbook(data: bytes) -> tuple[list[dict], list[dict]]:
    """
    Lee el .xlsx y devuelve (filas_validas, errores).

    filas_validas: [{"nombre":..., "identificador":..., "curso":..., "fila":n}]
    errores:       [{"fila": n, "motivo": "..."}]
    """
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # archivo corrupto o no es xlsx
        raise ValueError(f"No se pudo leer la planilla: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("La planilla está vacía.")

    # 1) Mapear encabezados de la primera fila no vacía.
    header_idx = next((i for i, r in enumerate(rows) if any(c is not None for c in r)), None)
    if header_idx is None:
        raise ValueError("La planilla está vacía.")

    header = rows[header_idx]
    colmap: dict[str, int] = {}
    for i, cell in enumerate(header):
        canon = _ALIAS.get(_norm(cell))
        if canon and canon not in colmap:
            colmap[canon] = i

    faltantes = [c for c in ("nombre", "identificador", "curso") if c not in colmap]
    if faltantes:
        legible = {"identificador": "rut"}
        nombres = ", ".join(legible.get(f, f) for f in faltantes)
        raise ValueError(
            f"Faltan las columnas: {nombres}. "
            "La primera fila debe tener los encabezados: nombre, rut, curso."
        )

    # 2) Leer las filas de datos.
    validas: list[dict] = []
    errores: list[dict] = []
    for offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue  # fila totalmente vacía: se ignora en silencio

        def val(key: str) -> str:
            i = colmap[key]
            return str(row[i]).strip() if i < len(row) and row[i] is not None else ""

        nombre, ident, curso = val("nombre"), val("identificador"), val("curso")

        if not nombre or not ident or not curso:
            faltan = [
                etiqueta
                for etiqueta, v in (("nombre", nombre), ("rut", ident), ("curso", curso))
                if not v
            ]
            errores.append({"fila": offset, "motivo": f"Falta {', '.join(faltan)}"})
            continue
        if len(nombre) < 2:
            errores.append({"fila": offset, "motivo": "El nombre es demasiado corto"})
            continue

        validas.append({"nombre": nombre, "identificador": ident, "curso": curso, "fila": offset})

    return validas, errores


def build_template_workbook() -> bytes:
    """Genera la planilla modelo que el profesor descarga y rellena."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    ws.append(["nombre", "rut", "curso"])
    ws.append(["Juan Pérez", "21.345.678-9", "1° Medio A"])
    ws.append(["Ana Rojas", "22.111.222-3", "1° Medio A"])
    ws.append(["Diego Muñoz", "23.444.555-6", "2° Medio B"])
    # Anchos cómodos para que se lea al abrirla.
    for col, width in (("A", 32), ("B", 18), ("C", 18)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
